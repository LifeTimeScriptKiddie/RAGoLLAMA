import os
import time
import logging
import uuid
import io
from datetime import datetime
from typing import List, Dict
import requests
import boto3
from botocore.exceptions import ClientError
import psycopg2
from psycopg2.extras import RealDictCursor
from qdrant_client import QdrantClient
from qdrant_client.models import Distance, VectorParams, PointStruct
import ollama
import PyPDF2
from docx import Document as DocxDocument
import openpyxl

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Environment variables
POSTGRES_USER = os.getenv("POSTGRES_USER", "rattg_user")
POSTGRES_PASSWORD = os.getenv("POSTGRES_PASSWORD", "changeme")
POSTGRES_HOST = os.getenv("POSTGRES_HOST", "localhost")
POSTGRES_DB = os.getenv("POSTGRES_DB", "rattgllm")

MINIO_ENDPOINT = os.getenv("MINIO_ENDPOINT", "localhost:9000")
MINIO_ACCESS_KEY = os.getenv("MINIO_ACCESS_KEY", "minioadmin")
MINIO_SECRET_KEY = os.getenv("MINIO_SECRET_KEY", "minioadmin")

QDRANT_HOST = os.getenv("QDRANT_HOST", "localhost")
QDRANT_PORT = int(os.getenv("QDRANT_PORT", "6333"))

OLLAMA_HOST = os.getenv("OLLAMA_HOST", "localhost")
OLLAMA_PORT = int(os.getenv("OLLAMA_PORT", "11434"))

BUCKET_NAME = "documents"
COLLECTION_NAME = "documents"
EMBEDDING_MODEL = "nomic-embed-text"
VECTOR_SIZE = 768
CHUNK_SIZE = 1000
CHUNK_OVERLAP = 200

class DocumentWorker:
    def __init__(self):
        self.setup_clients()
        
    def setup_clients(self):
        # Database connection
        self.db_conn_str = f"postgresql://{POSTGRES_USER}:{POSTGRES_PASSWORD}@{POSTGRES_HOST}/{POSTGRES_DB}"
        
        # MinIO client
        self.s3_client = boto3.client(
            's3',
            endpoint_url=f"http://{MINIO_ENDPOINT}",
            aws_access_key_id=MINIO_ACCESS_KEY,
            aws_secret_access_key=MINIO_SECRET_KEY,
            region_name='us-east-1'
        )
        
        # Qdrant client
        self.qdrant_client = QdrantClient(host=QDRANT_HOST, port=QDRANT_PORT)
        self.setup_qdrant_collection()
        
        # Ollama client
        self.ollama_client = ollama.Client(host=f"http://{OLLAMA_HOST}:{OLLAMA_PORT}")
        
    def setup_qdrant_collection(self):
        try:
            # Check if collection exists, create if it doesn't
            collections = self.qdrant_client.get_collections()
            collection_names = [col.name for col in collections.collections]
            
            if COLLECTION_NAME not in collection_names:
                self.qdrant_client.create_collection(
                    collection_name=COLLECTION_NAME,
                    vectors_config=VectorParams(size=VECTOR_SIZE, distance=Distance.COSINE),
                )
                logger.info(f"Created Qdrant collection: {COLLECTION_NAME}")
        except Exception as e:
            logger.error(f"Error setting up Qdrant collection: {e}")
            
    def get_db_connection(self):
        return psycopg2.connect(self.db_conn_str)
        
    def get_pending_documents(self):
        try:
            with self.get_db_connection() as conn:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    cur.execute("""
                        SELECT id, filename, original_filename, s3_key, user_id
                        FROM documents 
                        WHERE status = 'pending'
                        ORDER BY created_at ASC
                        LIMIT 10
                    """)
                    return cur.fetchall()
        except Exception as e:
            logger.error(f"Error fetching pending documents: {e}")
            return []
            
    def download_file(self, s3_key: str) -> bytes:
        try:
            response = self.s3_client.get_object(Bucket=BUCKET_NAME, Key=s3_key)
            return response['Body'].read()
        except ClientError as e:
            logger.error(f"Error downloading file {s3_key}: {e}")
            return None
            
    def extract_text(self, file_content: bytes, filename: str) -> str:
        try:
            file_extension = filename.split('.')[-1].lower()
            
            if file_extension == 'pdf':
                return self.extract_pdf_text(file_content)
            elif file_extension in ['docx', 'doc']:
                return self.extract_docx_text(file_content)
            elif file_extension in ['xlsx', 'xls']:
                return self.extract_excel_text(file_content)
            elif file_extension == 'txt':
                return file_content.decode('utf-8', errors='ignore')
            else:
                # Try to decode as text
                return file_content.decode('utf-8', errors='ignore')
                
        except Exception as e:
            logger.error(f"Error extracting text from {filename}: {e}")
            return ""
            
    def extract_pdf_text(self, file_content: bytes) -> str:
        try:
            pdf_file = io.BytesIO(file_content)
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text()
            return text
        except Exception as e:
            logger.error(f"Error extracting PDF text: {e}")
            return ""
            
    def extract_docx_text(self, file_content: bytes) -> str:
        try:
            doc = DocxDocument(io.BytesIO(file_content))
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        except Exception as e:
            logger.error(f"Error extracting DOCX text: {e}")
            return ""
            
    def extract_excel_text(self, file_content: bytes) -> str:
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    row_text = [str(cell) if cell is not None else "" for cell in row]
                    text += " ".join(row_text) + "\n"
            return text
        except Exception as e:
            logger.error(f"Error extracting Excel text: {e}")
            return ""
            
    def chunk_text(self, text: str) -> List[str]:
        chunks = []
        words = text.split()
        
        for i in range(0, len(words), CHUNK_SIZE - CHUNK_OVERLAP):
            chunk = " ".join(words[i:i + CHUNK_SIZE])
            chunks.append(chunk)
            
        return chunks
        
    def generate_embedding(self, text: str) -> List[float]:
        try:
            response = self.ollama_client.embeddings(
                model=EMBEDDING_MODEL,
                prompt=text
            )
            return response['embedding']
        except Exception as e:
            logger.error(f"Error generating embedding: {e}")
            return None
            
    def store_embeddings(self, document_id: int, chunks: List[str], filename: str):
        points = []
        embedding_records = []
        
        for i, chunk in enumerate(chunks):
            embedding = self.generate_embedding(chunk)
            if embedding is None:
                continue
                
            point_id = str(uuid.uuid4())
            
            # Qdrant point
            points.append(PointStruct(
                id=point_id,
                vector=embedding,
                payload={
                    "document_id": document_id,
                    "chunk_index": i,
                    "content": chunk,
                    "filename": filename
                }
            ))
            
            # Database record
            embedding_records.append((document_id, i, chunk, point_id))
            
        # Store in Qdrant
        if points:
            try:
                self.qdrant_client.upsert(
                    collection_name=COLLECTION_NAME,
                    points=points
                )
                logger.info(f"Stored {len(points)} embeddings in Qdrant for document {document_id}")
            except Exception as e:
                logger.error(f"Error storing embeddings in Qdrant: {e}")
                return False
                
        # Store in database
        try:
            with self.get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.executemany("""
                        INSERT INTO embeddings (document_id, chunk_index, content, vector_id, created_at)
                        VALUES (%s, %s, %s, %s, NOW())
                    """, embedding_records)
                    conn.commit()
                    logger.info(f"Stored {len(embedding_records)} embedding records in database")
        except Exception as e:
            logger.error(f"Error storing embeddings in database: {e}")
            return False
            
        return True
        
    def update_document_status(self, document_id: int, status: str, error_message: str = None):
        try:
            with self.get_db_connection() as conn:
                with conn.cursor() as cur:
                    if status == 'completed':
                        cur.execute("""
                            UPDATE documents 
                            SET status = %s, processed_at = NOW(), error_message = %s
                            WHERE id = %s
                        """, (status, error_message, document_id))
                    else:
                        cur.execute("""
                            UPDATE documents 
                            SET status = %s, error_message = %s
                            WHERE id = %s
                        """, (status, error_message, document_id))
                    conn.commit()
        except Exception as e:
            logger.error(f"Error updating document status: {e}")
            
    def process_document(self, document):
        document_id = document['id']
        s3_key = document['s3_key']
        filename = document['original_filename']
        
        logger.info(f"Processing document {document_id}: {filename}")
        
        try:
            # Update status to processing
            self.update_document_status(document_id, 'processing')
            
            # Download file
            file_content = self.download_file(s3_key)
            if file_content is None:
                self.update_document_status(document_id, 'failed', 'Failed to download file')
                return
                
            # Extract text
            text = self.extract_text(file_content, filename)
            if not text.strip():
                self.update_document_status(document_id, 'failed', 'No text could be extracted')
                return
                
            # Chunk text
            chunks = self.chunk_text(text)
            if not chunks:
                self.update_document_status(document_id, 'failed', 'No chunks generated')
                return
                
            # Generate and store embeddings
            if self.store_embeddings(document_id, chunks, filename):
                self.update_document_status(document_id, 'completed')
                logger.info(f"Successfully processed document {document_id}")
            else:
                self.update_document_status(document_id, 'failed', 'Failed to store embeddings')
                
        except Exception as e:
            logger.error(f"Error processing document {document_id}: {e}")
            self.update_document_status(document_id, 'failed', str(e))
            
    def run(self):
        logger.info("Document worker started")
        
        while True:
            try:
                # Get pending documents
                pending_docs = self.get_pending_documents()
                
                if pending_docs:
                    logger.info(f"Found {len(pending_docs)} pending documents")
                    for document in pending_docs:
                        self.process_document(document)
                else:
                    logger.info("No pending documents found")
                
                # Wait for 2 hours (7200 seconds) before next check
                logger.info("Waiting 2 hours before next check...")
                time.sleep(7200)
                
            except KeyboardInterrupt:
                logger.info("Worker stopped by user")
                break
            except Exception as e:
                logger.error(f"Error in worker main loop: {e}")
                time.sleep(60)  # Wait 1 minute before retrying

if __name__ == "__main__":
    worker = DocumentWorker()
    worker.run()